# 
# Cia. Porto Seguro - Coordenação TI Saúde
# Automatização Zeladoria (Sustentação)
# Gestão chamados do grupo Saude no PortoSDM
# PortoSDM é uma implementação do CA Desk Manager 
# By Daniel Faria Muniz - Outubro de 2018
#
# Baixa chamados (Incidente, Solicitação ou Problema) atribuídos ao grupo,
# gerando em seguida uma Pasta de Trabalho com uma Planilha para cada chamado,
# contendo seus cabeçalhos e respectivos históricos

import os
import sys
import subprocess
import shutil

PastaPython = os.path.dirname(sys.executable)
PastaScriptsPython = PastaPython + '\\Scripts\\' 

try:
    import time
except ImportError:
    print('time não encontrado; importando do PyPI...')
    print(subprocess.run([PastaScriptsPython + 'pip.exe', 'install', 'time']))
    import time

AnoMesDia = time.strftime('%Y%m%d')
AnoMesDiaHoraMinutoSegundo = time.strftime('%Y%m%d%H%M%S')

PastaRaiz = os.getcwd() + '\\'
PastaConfig = PastaRaiz + 'config\\'
PastaDrivers = PastaRaiz + 'drivers\\'
PastaDownloads = PastaRaiz + 'downloads\\' + AnoMesDia + '\\'

ConsultaChamado = '/CAisd/pdmweb.exe?OP=SEARCH+FACTORY=cr+SKIPLIST=1+QBE.EQ.ref_num='
ConsultaMudanca = '/CAisd/pdmweb.exe?OP=SEARCH+FACTORY=chg+SKIPLIST=1+QBE.IN.chg_ref_num='

try:
    import selenium
except ImportError:
    print('selenium não encontrado; importando do PyPI...')
    print(subprocess.run([PastaScriptsPython + 'pip.exe', 'install', 'selenium']))
    import selenium

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options

import winsound
import json

try:
    import glob
except ImportError:
    print('glob não encontrado; importando do PyPI...')
    print(subprocess.run([PastaScriptsPython + 'pip.exe', 'install', 'glob']))
    import glob

try:
    import sys
except ImportError:
    print('sys não encontrado; importando do PyPI...')
    print(subprocess.run([PastaScriptsPython + 'pip.exe', 'install', 'sys']))
    import sys

try:
    import bs4
except ImportError:
    print('beautifulsoup4 + lxml não encontrados; importando do PyPI...')
    print(subprocess.run([PastaScriptsPython + 'pip.exe', 'install', 'beautifulsoup4']))
    print(subprocess.run([PastaScriptsPython + 'pip.exe', 'install', 'lxml']))
    import bs4
from bs4 import BeautifulSoup

try:
    import openpyxl
except ImportError:
    print('openpyxl não encontrado; importando do PyPI...')
    print(subprocess.run([PastaScriptsPython + 'pip.exe', 'install', 'openpyxl']))
    import openpyxl

try:
    import pyautogui
except ImportError:
    print('pyautogui não encontrado; importando do PyPI...')
    print(subprocess.run([PastaScriptsPython + 'pip.exe', 'install', 'pyautogui']))
    import pyautogui
from pyautogui import press, typewrite, hotkey



def Clicar(BotaoAClicar):
    Botoes = Navegador.find_elements_by_xpath(BotaoAClicar)
    for Botao in Botoes:
        if (list(Botao.location.values())[0] != 0):
            Botao.click()
            break

def AbrePortoSDM(URLCompleta):
    print('PortoSDM: Abrindo e aguardando responder...', end = '')
    Navegador.get(URLCompleta)
    time.sleep(1)
    print('Ok!')

def BaixaChamados(Nome, Botao1, Botao2):
    Botao1 = '//span[text()="' + Botao1 + '"]'
    Botao2 = '//span[text()="' + Botao2 + '"]'
    Acao   = '//span[text()="Exportar("]'

    print('Navegação: ' + Nome + '...', end = '') 
    time.sleep(1)
    Navegador.switch_to.default_content()
    time.sleep(1)
    Navegador.switch_to.frame('product')
    time.sleep(1)
    Navegador.switch_to.frame('tab_2001')
    time.sleep(1)
    Navegador.switch_to.frame('role_main')
    time.sleep(1)
    Navegador.switch_to.frame('scoreboard')
    time.sleep(1)
    #Navegador.find_element_by_xpath(Botao1).click()

    Clicar(Botao1)

    print('Selecionando somente o que pertence ao grupo...', end = '')
    Navegador.find_element_by_xpath(Botao2).click()
    print('Ok!')

    print('Iniciando Exportação...', end = '')
    Navegador.switch_to_default_content()
    Navegador.switch_to.frame('product')
    time.sleep(1)
    Navegador.switch_to.frame('tab_2001')
    time.sleep(1)
    Navegador.switch_to.frame('role_main')
    time.sleep(1)
    Navegador.switch_to.frame('cai_main')
    time.sleep(1)
    Navegador.find_element_by_xpath(Acao).click()
    time.sleep(8)
    print('Concluindo Exportação...', end = '')
    time.sleep(8)
    print('Renomeando...', end = '') 
    os.rename(PastaDownloads + 'export.xls', PastaDownloads + AnoMesDiaHoraMinutoSegundo + Nome + '.xls')
    time.sleep(1)
    print('Ok!')

def PreenchePlanilha(Nome, PrefixoCabecalho, Planilha, CabecalhosDados, CelulasDados, LinhaInicial, ColunaInicial):
    NumColunas = 0
    TextoLinha = ''
    #
    #Estudo
    #------
    #Chamados
    # Incidentes
    #  BB1 'Mudança para correção   '
    #  BC1 'Causado pela requisição de mudança   '
    # Solicitações
    #  AQ1 'Atendimento através da Mudança   '
    #  AR1 'Causado pela requisição de mudança   '
    # Problemas
    #  M1 'Mudança   '
    #  N1 'Causado pela requisição de mudança   '
    #
    #Exemplos
    #--------
    #Incidente (BB e BC)
    #http://portosdm/CAisd/pdmweb.exe?OP=SEARCH+FACTORY=cr+SKIPLIST=1+QBE.EQ.ref_num=2885841
    #Solicitação (AQ e AR)
    #http://portosdm/CAisd/pdmweb.exe?OP=SEARCH+FACTORY=cr+SKIPLIST=1+QBE.EQ.ref_num=2826808
    #Problema (M e N)
    #http://portosdm/CAisd/pdmweb.exe?OP=SEARCH+FACTORY=cr+SKIPLIST=1+QBE.EQ.ref_num=1366073
    #
    ColunaMudancaSolucionadora = 0
    ColunaMudancaCausadora = 0

    for CabecalhoDados in CabecalhosDados:
        if CabecalhoDados.text != ' ':
            Texto = CabecalhoDados.text
            # Grava na Planilha
            Celula = Planilha.cell(row = LinhaInicial, column = ColunaInicial + NumColunas)     
            Celula.value = PrefixoCabecalho + '.' + Texto

            # Controle de RDMs
            if Texto in ['Mudança para correção   ', 'Atendimento através da Mudança   ',  'Mudança   ']:
                ColunaMudancaSolucionadora = ColunaInicial + NumColunas
                ColunaMudancaCausadora = ColunaInicial + NumColunas + 1
                print('Cabeçalho Mudança Solucionadora', ColunaMudancaSolucionadora)
                print('Cabeçalho Mudança Causadora', ColunaMudancaCausadora)

            TextoLinha = TextoLinha + Texto + '|'
            NumColunas = NumColunas + 1
    # Imprime no Terminal
    print(TextoLinha[:-1])

    Linha = 0
    Coluna = 0
    MaxRegistros = 20
    for CelulaDados in CelulasDados:

        # célula sem quebra de linha
        for e in CelulaDados.findAll('\n'):
            e.extract()
        # texto sem quebra de linha
        Texto = CelulaDados.getText().replace('\n',' ')

        # Grava na Planilha
        Celula = Planilha.cell(row = Linha + LinhaInicial + 1, column = Coluna + ColunaInicial)
        Celula.value = Texto


        # Controle de RDMs
        if (Linha == 0):
            if (Coluna + ColunaInicial) == (ColunaMudancaSolucionadora):
                if  (Texto != '   '):
                    print('Dados Mudança Solucionadora',Texto)
                    MudancasSolucionadoras.append(Texto.strip())
            if (Coluna + ColunaInicial) == (ColunaMudancaCausadora):
                if  (Texto != '   '):
                    print('Dados Mudança Causadora',Texto)               
                    MudancasCausadoras.append(Texto.strip())

        Coluna = Coluna + 1

        if (Coluna == NumColunas): # ultima coluna
            Final = '\n'    # quebra de linha
            Linha = Linha + 1
            Coluna = 0
        else:
            Final = '|'     # separador '|' (pipe)
        
        # Imprime no Terminal
        print(Texto, end=Final)

        if (Linha == MaxRegistros):
            break

    print('\n')   # quebra final de linha

    duration = 100  # milissegundos
    freq = 220  # Hz
    winsound.Beep(freq, duration)
    
    return ColunaInicial + NumColunas

def CarregaInformacoesChamado(Nome, PrefixoCabecalho, NumeroChamado):

    print('Carregando Informações do Chamado...', end = '')

    if (Servidor == "portosdm"):
        URLCompleta = 'http://' + Usuario + ':' + Senha + '@' + Servidor + ConsultaChamado + NumeroChamado
    else:
        URLCompleta = 'http://' + Usuario + ':' + Senha + '@' + Servidor + Porta + ConsultaChamado + NumeroChamado
        

    AbrePortoSDM(URLCompleta)

    Planilha = PastaTrabalhoHistoricoChamados[NumeroChamado]

    print('Hora de Web Scrapping!')

    # +--------------------------------------------------------------------------+
    # |                                 CABEÇALHO                                |
    # +--------------------------------------------------------------------------+
    print('Web Scrapping -> Cabeçalho')

    print('Contextualizando no Browser...', end = '')
    Navegador.switch_to.default_content()
    time.sleep(1)
    Navegador.switch_to.frame('product')
    time.sleep(1)
    Navegador.switch_to.frame('role_main')
    time.sleep(1)
    Navegador.find_element_by_id('acc_table_toplevel').click()
    time.sleep(1)
    print('Clicando em [4.Relacionamentos]...', end = '')
    Navegador.find_element_by_id('accrdnHyprlnk4').click()
    time.sleep(1)
    print('Clicando em ["2.Pai/filho]...', end = '')
    Navegador.find_element_by_id('tabHyprlnk4_2').click()
    time.sleep(1)
    print('Ok!')

    print('Preparando a Sopa [CABEÇALHO]...', end = '') 
    Sopa = BeautifulSoup(Navegador.page_source, 'lxml')

    print('Filtrando as Tabelas...', end = '') 
    Tabelas = Sopa.findAll('table',{'class':['tab', 'detailro']})

    LinhaInicial = 1
    ColunaInicial = 1

    for Tabela in Tabelas:

        IdTabela = str(Tabela.get('id'))

        if IdTabela.__contains__('dtltbl'):

            print('Extraindo a tabela', IdTabela, '...', end = '') 

            if Tabela is None:
                Texto = 'Lamento, mas não há ingrediente para fazer a sopa a partir da tabela' + IdTabela + '!'
                Celula = Planilha.cell(row = LinhaInicial, column = ColunaInicial)     
                Celula.value = Texto
                print(Texto)

                duration = 100  # milissegundos
                freq = 330  # Hz
                winsound.Beep(freq, duration)

                return

            print('Filtrando os Cabeçalhos dos Dados...', end = '')
            CabecalhosDados = Tabela.find_all(['tr','th'],attrs={'class':'detailro'})
            print('Filtrando as Celulas de Dados...', end = '')
            CelulasDados = Tabela.find_all(['tr','td'],attrs={'class':'detailro'})
            print('Ok!')

            ColunaInicial = PreenchePlanilha(Nome, PrefixoCabecalho, Planilha, CabecalhosDados, CelulasDados, LinhaInicial, ColunaInicial)
            duration = 100  # milissegundos
            freq = 220  # Hz
            winsound.Beep(freq, duration)

    # +--------------------------------------------------------------------------+
    # |                                 HISTORICO                                |
    # +--------------------------------------------------------------------------+
    print('Web Scrapping -> Histórico')

    print('Contextualizando no Browser...', end = '')
    Navegador.switch_to.default_content()
    time.sleep(1)
    Navegador.switch_to.frame('product')
    time.sleep(1)
    Navegador.switch_to.frame('role_main')
    time.sleep(1)
    Navegador.find_element_by_id('acc_table_toplevel').click()
    time.sleep(1)
    print('Clicando em [2.Logs]...', end = '')
    Navegador.find_element_by_id('accrdnHyprlnk2').click()
    time.sleep(1)
    print('Clicando em [1.Atividades]...', end = '')
    Navegador.find_element_by_id('tabHyprlnk2_1').click()
    time.sleep(1)
    print('Movendo para o Quadro...', end = '')
    Navegador.switch_to.frame('alg_iframe')
    time.sleep(1)
    print('Ok!')

    print('Preparando a Sopa [HISTÓRICO]...', end = '')
    Sopa = BeautifulSoup(Navegador.page_source, 'lxml')

    print('Filtrando a Tabela...', end = '')
    Tabela = Sopa.find('table',{'class':'ui-jqgrid-btable'})

    IdTabela = str(Tabela.get('id'))

    if Tabela is None:
        Texto = 'Lamento, mas não há ingrediente para fazer a sopa a partir da tabela' + IdTabela + '!'
        ColunaInicial = 1
        Celula = Planilha.cell(row = LinhaInicial, column = ColunaInicial)     
        Celula.value = Texto
        print(Texto)

        duration = 100  # milissegundos
        freq = 330  # Hz
        winsound.Beep(freq, duration)

        return

    print('Filtrando os Cabeçalhos dos Dados...', end = '')
    CabecalhosDados = Tabela.find_all(['a','span'],attrs={'class':'table_column_header_text'})
    print('Filtrando as Celulas de Dados...', end = '')
    CelulasDados = Tabela.find_all('td',{'style':'text-align:left;'})
    print('Ok!')

    LinhaInicial = LinhaInicial + 3
    ColunaInicial = 1
    PreenchePlanilha(Nome, PrefixoCabecalho, Planilha, CabecalhosDados, CelulasDados, LinhaInicial, ColunaInicial)

    duration = 100  # milissegundos
    freq = 220  # Hz
    winsound.Beep(freq, duration)

def CarregaInformacoesChamados(Nome, PrefixoCabecalho):
    print('Carregando informações de ' + Nome + '...')
    ArquivoEntrada = PastaDownloads + AnoMesDiaHoraMinutoSegundo + Nome + '.xls'
    Arquivo = open(ArquivoEntrada,'r', encoding='utf-8')
    Conteudo = Arquivo.read()
    Sopa = BeautifulSoup(Conteudo,'xml')
    Chamados = Sopa.find_all('Cell',attrs={'ss:StyleID':'string_url'})

    #Limitador (para propósito de debug)
    #MaxChamados = 2
    c = 0

    for Texto in Chamados:
        c = c + 1
        NumeroChamado = Texto.text.replace('\n','')
        print ('Processando [', NumeroChamado, '] -', c, 'de', len(Chamados), '...')

        if 1==1: #NumeroChamado in {'3133625','3133244','3133886','3133193','3126521','3119020'}
            print("Criando Guia '" + NumeroChamado + "'...")
            PastaTrabalhoHistoricoChamados.create_sheet(NumeroChamado)
            CarregaInformacoesChamado(Nome, PrefixoCabecalho, NumeroChamado)
            print('Atualizando Pasta de Trabalho...')
            PastaTrabalhoHistoricoChamados.save(ArquivoHistoricoChamados)

            #Limitador (para propósito de debug)
            #if c == MaxChamados:
            #    break

def CarregaInformacoesMudanca(Nome, PrefixoCabecalho, NumeroMudanca, PlanilhaMudanca):
    
    print('Carregando Informações da Mudança...', end = '')

    if (Servidor == "portosdm"):
        URLCompleta = 'http://' + Usuario + ':' + Senha + '@' + Servidor + ConsultaMudanca + NumeroMudanca
    else:
        URLCompleta = 'http://' + Usuario + ':' + Senha + '@' + Servidor + Porta + ConsultaMudanca + NumeroMudanca

    AbrePortoSDM(URLCompleta)

    print('Hora de Web Scrapping!')

    # +--------------------------------------------------------------------------+
    # |                                 CABEÇALHO                                |
    # +--------------------------------------------------------------------------+
    print('Web Scrapping -> Cabeçalho')

    print('Contextualizando no Browser...', end = '')
    Navegador.switch_to.default_content()
    time.sleep(1)
    Navegador.switch_to.frame('product')
    time.sleep(1)
    Navegador.switch_to.frame('role_main')
    time.sleep(1)
    Navegador.find_element_by_id('acc_table_toplevel').click()
    time.sleep(1)
    print('Ok!')

    print('Preparando a Sopa [CABEÇALHO]...', end = '') 
    Sopa = BeautifulSoup(Navegador.page_source, 'lxml')

    print('Filtrando as Tabelas...', end = '') 
    Tabelas = Sopa.findAll('table',{'class':['tab', 'detailro']})

    LinhaInicial = 1
    ColunaInicial = 1

    for Tabela in Tabelas:

        IdTabela = str(Tabela.get('id'))

        if IdTabela.__contains__('dtltbl'):

            print('Extraindo a tabela', IdTabela, '...', end = '') 

            if Tabela is None:
                Texto = 'Lamento, mas não há ingrediente para fazer a sopa a partir da tabela' + IdTabela + '!'
                Celula = PlanilhaMudanca.cell(row = LinhaInicial, column = ColunaInicial)     
                Celula.value = Texto
                print(Texto)

                duration = 100  # milissegundos
                freq = 330  # Hz
                winsound.Beep(freq, duration)

                return

            print('Filtrando os Cabeçalhos dos Dados...', end = '')
            CabecalhosDados = Tabela.find_all(['tr','th'],attrs={'class':'detailro'})
            print('Filtrando as Celulas de Dados...', end = '')
            CelulasDados = Tabela.find_all(['tr','td'],attrs={'class':'detailro'})
            print('Ok!')

            ColunaInicial = PreenchePlanilha(Nome, PrefixoCabecalho, PlanilhaMudanca, CabecalhosDados, CelulasDados, LinhaInicial, ColunaInicial)
            duration = 100  # milissegundos
            freq = 220  # Hz
            winsound.Beep(freq, duration)

    # +--------------------------------------------------------------------------+
    # |                                 CHAMADOS RELACIONADOS                    |
    # +--------------------------------------------------------------------------+
    print('Web Scrapping -> Chamados Relacionados')

    print('Contextualizando no Browser...', end = '')
    Navegador.switch_to.default_content()
    time.sleep(1)
    Navegador.switch_to.frame('product')
    time.sleep(1)
    Navegador.switch_to.frame('role_main')
    time.sleep(1)
    Navegador.find_element_by_id('acc_table_toplevel').click()
    time.sleep(1)
    print('Clicando em [3.Tickets Relacionados]...', end = '')
    Navegador.find_element_by_id('accrdnHyprlnk3').click()
    time.sleep(1)
    print('Clicando em ["2.Incidentes/Problemas]...', end = '')
    Navegador.find_element_by_id('tabHyprlnk3_2').click()
    time.sleep(1)
    print('Ok!')
    print('Movendo para o Quadro...', end = '')
    Navegador.switch_to.frame('rel_iframe')
    time.sleep(1)
    print('Ok!')

    print('Preparando a Sopa [HISTÓRICO]...', end = '')
    Sopa = BeautifulSoup(Navegador.page_source, 'lxml')

    print('Filtrando a Tabela...', end = '')
    Tabela = Sopa.find('table',{'class':'ui-jqgrid-btable'})

    IdTabela = str(Tabela.get('id'))

    if Tabela is None:
        Texto = 'Lamento, mas não há ingrediente para fazer a sopa a partir da tabela' + IdTabela + '!'
        ColunaInicial = 1
        Celula = PlanilhaMudanca.cell(row = LinhaInicial, column = ColunaInicial)     
        Celula.value = Texto
        print(Texto)

        duration = 100  # milissegundos
        freq = 330  # Hz
        winsound.Beep(freq, duration)

        return

    print('Filtrando os Cabeçalhos dos Dados...', end = '')
    CabecalhosDados = Tabela.find_all(['a','span'],attrs={'class':'table_column_header_text'})
    # PAREI AQUI - 2018.10.24 10:25
    #CabecalhosDados = Tabela.find_all(['th','span'],attrs={'style':'','class':'table_column_header_text'})
    #CabecalhosDados = Tabela.find_all(['a'],attrs={'class':'table_column_header_text'})
    #CabecalhosDados = Tabela.find_all(['th'],attrs={'style':''}).find_all(['a','span'],attrs={'class':'table_column_header_text'})
    print('Filtrando as Celulas de Dados...', end = '')
    CelulasDados = Tabela.find_all('td',{'style':'text-align:left;'})
    print('Ok!')

    LinhaInicial = LinhaInicial + 10
    ColunaInicial = 1
    PreenchePlanilha(Nome, PrefixoCabecalho, PlanilhaMudanca, CabecalhosDados, CelulasDados, LinhaInicial, ColunaInicial)

    duration = 100  # milissegundos
    freq = 220  # Hz
    winsound.Beep(freq, duration)

def CarregaInformacoesMudancas(Nome, PrefixoCabecalho, PastaTrabalhoMudanca, ArquivoMudanca, Mudancas):
    print('Carregando informações de ' + ArquivoMudanca + '...')

    c = 0

    for NumeroMudanca in Mudancas:

        c = c + 1
        print ('Processando [', NumeroMudanca, '] -', c, 'de', len(Mudancas), '...')

        if 1==1:
            if not (NumeroMudanca in PastaTrabalhoMudanca):
                print("Criando Guia '" + NumeroMudanca + "'...")
                PastaTrabalhoMudanca.create_sheet(NumeroMudanca)
                PlanilhaMudanca = PastaTrabalhoMudanca[NumeroMudanca]
                CarregaInformacoesMudanca(Nome, PrefixoCabecalho, NumeroMudanca, PlanilhaMudanca)
            print('Atualizando Pasta de Trabalho...')
            PastaTrabalhoMudanca.save(ArquivoMudanca)

#
# +---------+
# | M A I N |
# +---------+
#

print('+------------------------+')
print('| INICIANDO A AUTOMAÇÃO! |')
print('+------------------------+')
time.sleep(1)

print('Inicializando listas de Mudanças (RDMs) Solucionadoras e Causadoras...')
MudancasSolucionadoras = []
MudancasCausadoras = []

print('Lendo configuração...')
with open(PastaConfig + 'config.json') as ArquivoJson:  
    Dados = json.load(ArquivoJson)
    Usuario = Dados['usuario']
    print('\tUsuario: '+ (Usuario))
    Senha = Dados['senha']
    Servidor = Dados['servidor']
    print('\tServidor: '+ (Servidor))
    Porta = Dados['porta']
    print('\tPorta: '+ (Porta))
    PastaDestino = Dados['pastadestino']
    print('\tPorta: '+ (PastaDestino))

if (Servidor == "portosdm"):
    URLCompleta = 'http://' + Usuario + ':' + Senha + '@' + Servidor
else:
    URLCompleta = 'http://' + Usuario + ':' + Senha + '@' + Servidor + Porta

print("Verificando existência da pasta '" + PastaDownloads + "'...", end = '')
if os.path.exists(PastaDownloads):
    print('Já existe!')
else:
    print('Criando...', end = '')
    os.mkdir(PastaDownloads)
    print ('Ok!')

print('Verificando se há exportações antigas, provavelmente abortadas antes (export*.xls)...', end = '')
for file in glob.glob(PastaDownloads + 'export*.xls'):
    print("Apagando '" + file +"'...", end = '')
    os.remove(file)
print ('Ok!')

print('Chrome WebDriver: Preparando...')
time.sleep(30)
Opcoes = webdriver.ChromeOptions()
Preferencias = {'download.default_directory' : PastaDownloads}
Opcoes.add_experimental_option('prefs', Preferencias)
Opcoes.add_argument('disable-notifications')
# Opcoes.add_argument('--silent')
Opcoes.add_argument('--start-maximized')
# Opcoes.add_argument("--incognito")
# Opcoes.add_argument('C:\\Users\\1620300\\AppData\\Local\\Google\\Chrome\\User Data\\Default')
Navegador = webdriver.Chrome(options = Opcoes, executable_path = PastaDrivers + 'chromedriver.exe')
time.sleep(1)

AbrePortoSDM(URLCompleta)

BaixaChamados('Incidentes','Incidentes','Incidentes do meu grupo')
# BaixaChamados('Solicitacoes','Solicitação','Solicitações do meu grupo')
# BaixaChamados('Problemas','Problema','Problemas do meu grupo')

print("Verificando existência da pasta '" + PastaDestino + "'...", end = '')
if os.path.exists(PastaDestino):
    print('Já existe!')
else:
    print('Criando...', end = '')
    os.mkdir(PastaDestino)
    print ('Ok!')

print("Copiando arquivos da Origem '" + PastaDownloads + "' para o Destino '" + PastaDestino + "'...")
for Arquivo in glob.glob(PastaDownloads + AnoMesDiaHoraMinutoSegundo + '*.xls'):
    print(Arquivo)
    shutil.copy2(Arquivo, PastaDestino) # 'copy2' é praticamente idêntico ao 'copy()', com a diferença que também tenta preservar todos os metadados do arquivo (timestamp por exemplo).

#Pular este codigo, por enquanto
if (0==0):
    print("Criando Pasta de Trabalho dos históricos dos chamados...")
    PastaTrabalhoHistoricoChamados = openpyxl.Workbook()

    print("Salvando (versão inicial)...")
    ArquivoHistoricoChamados =  PastaDownloads + AnoMesDiaHoraMinutoSegundo + 'HistoricoChamados.xlsx'
    PastaTrabalhoHistoricoChamados.save(ArquivoHistoricoChamados)

    CarregaInformacoesChamados('Incidentes','Incidente')
    CarregaInformacoesChamados('Solicitacoes','Solicitacao')
    CarregaInformacoesChamados('Problemas','Problema')

    print('Salvando e Fechando Pasta de Trabalho...')
    Padrao = PastaTrabalhoHistoricoChamados['Sheet']
    PastaTrabalhoHistoricoChamados.remove(Padrao)
    PastaTrabalhoHistoricoChamados.save(ArquivoHistoricoChamados)
    PastaTrabalhoHistoricoChamados.close()

    Nome = 'MudancasSolucionadoras'
    PrefixoCabecalho = 'MudancaSolucionadora'
    print(Nome + ': ', MudancasSolucionadoras)
    if len(MudancasSolucionadoras) > 0:
        print('Criando Pasta de Trabalho das ' + Nome + '...')
        PastaTrabalhoMudancasSolucionadoras = openpyxl.Workbook()
        print('Salvando (versão inicial)...')
        ArquivoSaidaMudancasSolucionadoras =  PastaDownloads + AnoMesDiaHoraMinutoSegundo + Nome + '.xlsx'
        PastaTrabalhoMudancasSolucionadoras.save(ArquivoSaidaMudancasSolucionadoras)
        CarregaInformacoesMudancas(Nome, PrefixoCabecalho, PastaTrabalhoMudancasSolucionadoras, ArquivoSaidaMudancasSolucionadoras, MudancasSolucionadoras)
        print('Salvando e Fechando Pasta de Trabalho...')
        Padrao = PastaTrabalhoMudancasSolucionadoras['Sheet']
        PastaTrabalhoMudancasSolucionadoras.remove(Padrao)
        PastaTrabalhoMudancasSolucionadoras.save(ArquivoSaidaMudancasSolucionadoras)
        PastaTrabalhoMudancasSolucionadoras.close()

    Nome = 'MudancasCausadoras'
    PrefixoCabecalho = 'MudancaCausadora'
    print(Nome, ': ', MudancasCausadoras)
    if len(MudancasCausadoras) > 0:
        print('Criando Pasta de Trabalho das ' + Nome + '...')
        PastaTrabalhoMudancasCausadoras = openpyxl.Workbook()
        print('Salvando (versão inicial)...')
        ArquivoSaidaMudancasCausadoras =  PastaDownloads + AnoMesDiaHoraMinutoSegundo + Nome + '.xlsx'
        PastaTrabalhoMudancasCausadoras.save(ArquivoSaidaMudancasCausadoras)
        CarregaInformacoesMudancas(Nome, PrefixoCabecalho, PastaTrabalhoMudancasCausadoras, ArquivoSaidaMudancasCausadoras, MudancasCausadoras)
        print('Salvando e Fechando Pasta de Trabalho...')
        Padrao = PastaTrabalhoMudancasCausadoras['Sheet']
        PastaTrabalhoMudancasCausadoras.remove(Padrao)
        PastaTrabalhoMudancasCausadoras.save(ArquivoSaidaMudancasCausadoras)
        PastaTrabalhoMudancasCausadoras.close()

    print("Copiando arquivos da Origem '" + PastaDownloads + "' para o Destino '" + PastaDestino + "'...")
    for Arquivo in glob.glob(PastaDownloads + AnoMesDiaHoraMinutoSegundo + '*.xlsx'):
        print(Arquivo)
        shutil.copy2(Arquivo, PastaDestino) # 'copy2' é praticamente idêntico ao 'copy()', com a diferença que também tenta preservar todos os metadados do arquivo (timestamp por exemplo).

print('Chrome WebDriver: Fechando...')
time.sleep(1)
Navegador.quit()

print("Abrindo pasta '" + PastaDestino + "'...")
os.startfile(PastaDestino)
time.sleep(1)

''' print("Abrindo pasta '" + PastaDownloads + "'...")
os.startfile(PastaDownloads)
time.sleep(1) '''


duration = 1000  # milissegundos
freq = 110  # Hz
winsound.Beep(freq, duration)
time.sleep(1)

print('+------------+')
print('| CONCLUÍDO! |')
print('+------------+')
time.sleep(1)
sys.exit(0)
